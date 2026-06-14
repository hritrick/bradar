"""Report generation: JSON + PDF + Excel + CSV — SQL-aggregated for scale."""
import csv
import io
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
from models import Business, Prediction, LeadScore, DailyReport

log = logging.getLogger(__name__)
REPORTS_DIR = Path("/app/backend/reports_out")
REPORTS_DIR.mkdir(exist_ok=True, parents=True)


async def build_report_data(
    db: AsyncSession,
    *,
    report_date: Optional[date] = None,
    filters: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    report_date = report_date or date.today()
    day_start = datetime.combine(report_date, datetime.min.time())
    day_end = day_start + timedelta(days=1)

    total = (await db.execute(select(func.count(Business.id)))).scalar() or 0
    today_count = (await db.execute(
        select(func.count(Business.id)).where(Business.created_at >= day_start, Business.created_at < day_end)
    )).scalar() or 0

    # Latest LeadScore per business via SQL
    latest = (
        select(LeadScore.business_id, func.max(LeadScore.created_at).label("mx"))
        .group_by(LeadScore.business_id)
        .subquery()
    )
    latest_join = (
        select(LeadScore.business_id, LeadScore.score, LeadScore.lead_category)
        .join(latest, and_(LeadScore.business_id == latest.c.business_id, LeadScore.created_at == latest.c.mx))
    ).subquery()

    cat_counts = (await db.execute(
        select(latest_join.c.lead_category, func.count()).group_by(latest_join.c.lead_category)
    )).all()
    by_cat = {r[0]: r[1] for r in cat_counts}
    hot, warm, cold = int(by_cat.get("HOT") or 0), int(by_cat.get("WARM") or 0), int(by_cat.get("COLD") or 0)
    avg = float((await db.execute(select(func.avg(latest_join.c.score)))).scalar() or 0)

    by_city_rows = (await db.execute(
        select(Business.city, func.count(Business.id)).where(Business.city.isnot(None)).group_by(Business.city).order_by(func.count(Business.id).desc()).limit(15)
    )).all()
    by_industry_rows = (await db.execute(
        select(Business.industry, func.count(Business.id)).where(Business.industry.isnot(None)).group_by(Business.industry).order_by(func.count(Business.id).desc()).limit(15)
    )).all()
    by_cat_rows = (await db.execute(
        select(Business.category, func.count(Business.id)).where(Business.category.isnot(None)).group_by(Business.category).order_by(func.count(Business.id).desc()).limit(15)
    )).all()
    pred_rows = (await db.execute(
        select(Prediction.predicted_need, func.count(Prediction.id)).where(Prediction.predicted_need.isnot(None)).group_by(Prediction.predicted_need).order_by(func.count(Prediction.id).desc()).limit(15)
    )).all()
    pin_rows = (await db.execute(
        select(Business.pincode, func.count(Business.id)).where(Business.pincode.isnot(None)).group_by(Business.pincode).order_by(func.count(Business.id).desc()).limit(15)
    )).all()

    today_list_res = (await db.execute(
        select(Business).where(Business.created_at >= day_start, Business.created_at < day_end)
        .order_by(Business.created_at.desc()).limit(50)
    )).scalars().all()
    # Latest score for today_list
    today_ids = [b.id for b in today_list_res]
    score_map = {}
    if today_ids:
        rows = (await db.execute(select(latest_join).where(latest_join.c.business_id.in_(today_ids)))).all()
        score_map = {r[0]: {"score": r[1], "lead_category": r[2]} for r in rows}
    today_list = []
    for b in today_list_res:
        s = score_map.get(b.id) or {}
        today_list.append({
            "id": b.id, "business_name": b.business_name, "city": b.city,
            "industry": b.industry, "category": b.category,
            "director_name": b.director_name, "phone": b.phone, "email": b.email,
            "registration_date": str(b.registration_date) if b.registration_date else None,
            "score": s.get("score"), "lead_category": s.get("lead_category"),
        })

    return {
        "report_date": str(report_date),
        "generated_at": datetime.utcnow().isoformat(),
        "kpis": {
            "total_businesses": int(total),
            "today_new": int(today_count),
            "hot_leads": hot, "warm_leads": warm, "cold_leads": cold,
            "avg_score": round(avg, 1),
        },
        "by_city": [{"city": r[0], "count": r[1]} for r in by_city_rows],
        "by_industry": [{"industry": r[0], "count": r[1]} for r in by_industry_rows],
        "by_category": [{"category": r[0], "count": r[1]} for r in by_cat_rows],
        "by_predicted_need": [{"need": r[0], "count": r[1]} for r in pred_rows],
        "by_pincode": [{"pincode": r[0], "count": r[1]} for r in pin_rows],
        "today_list": today_list,
        "filters": filters or {},
    }


def generate_pdf(report: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Title"], textColor=colors.HexColor("#1A1A1A"))
    subtitle = ParagraphStyle("subtitle", parent=styles["Normal"], textColor=colors.HexColor("#6B6B6B"), fontSize=10)
    section = ParagraphStyle("section", parent=styles["Heading2"], textColor=colors.HexColor("#1A1A1A"))

    el = []
    el.append(Paragraph("Business Radar AI", title))
    el.append(Paragraph(f"Daily Intelligence Report — {report['report_date']}", subtitle))
    el.append(Spacer(1, 8))

    k = report["kpis"]
    kpi_data = [
        ["Total Businesses", "Today's New", "Hot", "Warm", "Cold", "Avg Score"],
        [k["total_businesses"], k["today_new"], k["hot_leads"], k["warm_leads"], k["cold_leads"], k["avg_score"]],
    ]
    kpi_table = Table(kpi_data, colWidths=[28 * mm] * 6)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F0F10")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E5E5")),
    ]))
    el.append(kpi_table)
    el.append(Spacer(1, 14))

    el.append(Paragraph("Top Industries", section))
    ind_data = [["Industry", "Count"]] + [[c.get("industry") or "—", c["count"]] for c in report.get("by_industry", [])[:8]]
    ind_table = Table(ind_data, colWidths=[100 * mm, 30 * mm])
    ind_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F0F10")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E5E5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F8")]),
    ]))
    el.append(ind_table)
    el.append(Spacer(1, 14))

    el.append(Paragraph("Today's New Businesses", section))
    rows = [["Business", "City", "Industry", "Score", "Lead"]]
    for b in report["today_list"][:20]:
        rows.append([b["business_name"], b["city"] or "—", (b.get("industry") or b.get("category") or "—"), b["score"] or 0, b["lead_category"] or "—"])
    tbl = Table(rows, colWidths=[60 * mm, 28 * mm, 35 * mm, 18 * mm, 24 * mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F0F10")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E5E5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F8")]),
    ]))
    el.append(tbl)
    el.append(Spacer(1, 18))
    el.append(Paragraph(f"Generated: {report['generated_at']} — Business Radar AI", subtitle))
    doc.build(el)
    return buf.getvalue()


def businesses_to_csv(rows: List[Dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    if not rows:
        return b""
    fieldnames = list(rows[0].keys())
    w = csv.DictWriter(buf, fieldnames=fieldnames)
    w.writeheader()
    for r in rows:
        w.writerow({k: (str(v) if v is not None else "") for k, v in r.items()})
    return buf.getvalue().encode("utf-8")


def businesses_to_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Businesses"
    if not rows:
        wb.save(io.BytesIO())
        return b""
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF0F0F10")
    for r in rows:
        ws.append([(str(v) if not isinstance(v, (int, float)) and v is not None else v) for v in r.values()])
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(40, max(12, length + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def fetch_businesses_for_export(
    db: AsyncSession,
    *,
    ids: Optional[List[str]] = None,
    filters: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    q = select(Business)
    if ids:
        q = q.where(Business.id.in_(ids))
    res = await db.execute(q.order_by(Business.created_at.desc()).limit(5000))
    bizs = res.scalars().all()
    if not bizs:
        return []
    biz_ids = [b.id for b in bizs]
    ls_res = await db.execute(select(LeadScore).where(LeadScore.business_id.in_(biz_ids)))
    by_biz_score = {}
    for ls in ls_res.scalars():
        cur = by_biz_score.get(ls.business_id)
        if not cur or (ls.created_at and (not cur.created_at or ls.created_at > cur.created_at)):
            by_biz_score[ls.business_id] = ls
    pr_res = await db.execute(select(Prediction).where(Prediction.business_id.in_(biz_ids)))
    by_biz_pred = {}
    for p in pr_res.scalars():
        cur = by_biz_pred.get(p.business_id)
        if not cur or (p.created_at and (not cur.created_at or p.created_at > cur.created_at)):
            by_biz_pred[p.business_id] = p
    out = []
    for b in bizs:
        ls = by_biz_score.get(b.id)
        pr = by_biz_pred.get(b.id)
        out.append({
            "id": b.id, "business_name": b.business_name, "gst_number": b.gst_number,
            "registration_date": str(b.registration_date) if b.registration_date else None,
            "company_type": b.company_type, "industry": b.industry,
            "category": b.category, "sub_category": b.sub_category,
            "director_name": b.director_name, "phone": b.phone, "email": b.email,
            "website": b.website, "linkedin_url": b.linkedin_url,
            "employee_estimate": b.employee_estimate,
            "address": b.address, "locality": b.locality, "city": b.city, "district": b.district,
            "state": b.state, "country": b.country, "pincode": b.pincode,
            "latitude": b.latitude, "longitude": b.longitude,
            "source": b.source, "confidence_score": b.confidence_score,
            "lead_score": ls.score if ls else None,
            "lead_category": ls.lead_category if ls else None,
            "predicted_need": pr.predicted_need if pr else None,
            "need_probability": pr.probability if pr else None,
        })
    return out
